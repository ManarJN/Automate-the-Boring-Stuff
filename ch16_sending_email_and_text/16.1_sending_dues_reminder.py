#! /usr/bin/env python3

# Automate the Boring Stuff
# Chapter 16 - Sending Emails and Texts
# Send Dues Reminders - Sends emails based on payment status in spreadsheet

import openpyxl
import smtplib
import sys

# open the spreadsheet and get the latest dues status
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

# check each member's payment status
unpaidMembers = {}
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# log in to email account
smptObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj = ehlo()
smtpObj.starttls()
smtpObj.login('email@gmail.com', sys.argv[1])

# send out reminder emails
for name, email in unpaidMembers.items()
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not " \
           "paid dues for %s. Please make this payment as soon as possible. Thank " \
           "you!" % (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('email@gmail.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

smtpObj.quit()