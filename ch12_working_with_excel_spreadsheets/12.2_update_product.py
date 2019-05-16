#! /usr/bin/env python3

# Automate the Boring Stuff
# Chapter 12 - Working with Excel Worksheets
# Update Produce - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('./12.2_files/input_excel/produce_sales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# loops through the rows and updates the prices
for rowNum in range(2, sheet.max_row):  #skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('./12.2_files/output_excel/produce_sales.xlsx')